import openpyxl


def getRowCount(file, sheetName):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    return sheet.max_row


def getColCount(file, sheetName):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    return sheet.max_column


def readData(file, sheetName, row, col):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    return sheet.cell(row=row, column=col).value